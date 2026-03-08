"""
LapseTable: actuarial lapse table in qx/px/lx format.

Actuaries recognise tables. The qx/px/lx format is the standard output from
life and general insurance persistence models. This class wraps any fitted
survival model and produces tables in that format.

Multi-decrement tables (lapse + cancel_mtd + total_loss as separate columns)
are not implemented in v0.1.0, but the structure is set up to support them.
"""

from __future__ import annotations

from typing import Any

import numpy as np
import polars as pl

from ._utils import to_polars


class LapseTable:
    """Actuarial multi-decrement lapse table from a fitted survival model.

    Generates tables in the qx/px/lx format actuaries use:

        lx  = number of lives at start of year x (radix = 10,000)
        dx  = deaths (lapses) in year x: lx - lx+1
        qx  = probability of exiting in year x: dx / lx
        px  = 1 - qx = probability of surviving year x
        Tx  = curtate expected future lifetime from year x

    Parameters
    ----------
    survival_model : Any
        Fitted lifelines fitter or WeibullMixtureCureFitter.
    radix : int
        Starting population for lx column. Default 10,000.
    time_points : list[float] | None
        Policy year boundaries at which to evaluate.
        Default [1, 2, 3, 4, 5, 6, 7].

    Examples
    --------
    >>> from insurance_survival import LapseTable
    >>>
    >>> table = LapseTable(survival_model=aft, radix=10_000)
    >>> df = table.generate(covariate_profile={"ncd_level": 3, "channel": "direct"})
    >>> print(df)
    shape: (7, 6)
    ┌──────┬───────┬──────┬───────┬───────┬───────┐
    │ year │ lx    │ dx   │ qx    │ px    │ Tx    │
    │ ---  │ ---   │ ---  │ ---   │ ---   │ ---   │
    │ i32  │ i32   │ i32  │ f64   │ f64   │ f64   │
    ╞══════╪═══════╪══════╪═══════╪═══════╪═══════╡
    │ 1    │ 10000 │ 2300 │ 0.230 │ 0.770 │ 2.84  │
    ...
    """

    def __init__(
        self,
        survival_model: Any,
        radix: int = 10_000,
        time_points: list[float] | None = None,
    ) -> None:
        self.survival_model = survival_model
        self.radix = radix
        self.time_points = time_points if time_points is not None else [1, 2, 3, 4, 5, 6, 7]

        self._last_table: pl.DataFrame | None = None

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def generate(
        self,
        covariate_profile: dict[str, Any] | pl.DataFrame,
        by: str | None = None,
    ) -> pl.DataFrame:
        """Generate lapse table for a given covariate profile.

        Parameters
        ----------
        covariate_profile : dict | pl.DataFrame
            Either a single dict (one table) or a DataFrame of profiles
            (concatenated with a segment identifier column).
        by : str | None
            If covariate_profile is a DataFrame, name of the column to use
            as segment label in output.

        Returns
        -------
        pl.DataFrame
            Columns: year, lx, dx, qx, px, Tx.
            Plus segment column if by is set.
        """
        if isinstance(covariate_profile, dict):
            profile_df = pl.DataFrame([covariate_profile])
            tables = [self._generate_single(profile_df)]
            self._last_table = tables[0]
            return tables[0]

        profile_df = to_polars(covariate_profile)

        if by is None or by not in profile_df.columns:
            # Single table using first row
            table = self._generate_single(profile_df.head(1))
            self._last_table = table
            return table

        # One table per segment
        tables = []
        for row in profile_df.iter_rows(named=True):
            row_df = pl.DataFrame([row])
            table = self._generate_single(row_df)
            segment_label = row[by]
            table = table.with_columns(
                pl.lit(segment_label).cast(pl.Utf8).alias(by)
            )
            tables.append(table)

        result = pl.concat(tables)
        self._last_table = result
        return result

    def to_excel(self, path: str) -> None:
        """Write the last-generated table to Excel (openpyxl).

        Parameters
        ----------
        path : str
            Output file path (must end in .xlsx).
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError:
            raise ImportError(
                "openpyxl is required for to_excel(). "
                "Install with: pip install insurance-survival[excel]"
            )

        if self._last_table is None:
            raise RuntimeError("Call generate() before to_excel().")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Lapse Table"

        pdf = self._last_table.to_pandas()
        for r_idx, row in enumerate(
            dataframe_to_rows(pdf, index=False, header=True), 1
        ):
            ws.append(row)
            if r_idx == 1:
                # Bold header
                for cell in ws[r_idx]:
                    cell.font = Font(bold=True)

        wb.save(path)

    # ------------------------------------------------------------------
    # Private helpers
    # ------------------------------------------------------------------

    def _get_survival_at_times(
        self, profile_df: pl.DataFrame, times: list[float]
    ) -> np.ndarray:
        """Get S(t) at each time point for the given covariate profile.

        Returns np.ndarray of shape (len(times),).
        """
        from .cure import WeibullMixtureCureFitter

        # Include t=0 for computing first qx
        all_times = [0.0] + list(times)

        if isinstance(self.survival_model, WeibullMixtureCureFitter):
            surv_df = self.survival_model.predict_survival_function(
                profile_df, times=all_times
            )
            # Columns: S_t1, S_t2, ... (index 0 = t=0)
            sf_values = np.array([
                float(surv_df[f"S_t{k + 1}"][0]) for k in range(len(all_times))
            ])
        else:
            # lifelines fitter
            pdf = profile_df.to_pandas()
            # Drop non-numeric string columns that lifelines cannot use
            object_cols = [c for c in pdf.columns if pdf[c].dtype == object]
            if object_cols:
                pdf = pdf.drop(columns=object_cols)
            sf = self.survival_model.predict_survival_function(
                pdf, times=all_times
            )
            sf_values = sf.iloc[:, 0].values  # (len(all_times),)

        return sf_values

    def _generate_single(self, profile_df: pl.DataFrame) -> pl.DataFrame:
        """Generate a lapse table for one covariate profile."""
        times = self.time_points
        sf_values = self._get_survival_at_times(profile_df, times)
        # sf_values[0] = S(0) = 1.0 (includes t=0 at index 0)
        # sf_values[1] = S(t_1), ...

        # Enforce monotone non-increasing survival (numerical noise from
        # optimiser can produce tiny inversions at adjacent time points).
        sf_values = np.minimum.accumulate(sf_values)

        # lx at each time point
        lx_float = np.array([sf_values[k + 1] for k in range(len(times))]) * self.radix
        lx_0 = self.radix  # starting cohort

        # Prepend l_0 for computing dx
        lx_with_start = np.concatenate([[lx_0], lx_float])

        years = []
        lx_vals = []
        dx_vals = []
        qx_vals = []
        px_vals = []
        Tx_vals = []

        for k in range(len(times)):
            l_curr = lx_with_start[k]
            l_next = lx_with_start[k + 1]

            dx = l_curr - l_next
            qx = dx / l_curr if l_curr > 0 else 0.0
            px = 1.0 - qx

            # Curtate expected future lifetime from year k+1:
            # Tx = sum_{j=k}^{T-1} lx_{j+1} / lx_k
            Tx = sum(lx_with_start[j + 1] for j in range(k, len(times))) / l_curr if l_curr > 0 else 0.0

            years.append(k + 1)
            lx_vals.append(int(round(l_curr)))
            dx_vals.append(int(round(dx)))
            qx_vals.append(round(float(qx), 6))
            px_vals.append(round(float(px), 6))
            Tx_vals.append(round(float(Tx), 4))

        return pl.DataFrame({
            "year": pl.Series(years, dtype=pl.Int32),
            "lx": pl.Series(lx_vals, dtype=pl.Int32),
            "dx": pl.Series(dx_vals, dtype=pl.Int32),
            "qx": pl.Series(qx_vals, dtype=pl.Float64),
            "px": pl.Series(px_vals, dtype=pl.Float64),
            "Tx": pl.Series(Tx_vals, dtype=pl.Float64),
        })
